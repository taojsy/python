"""
collection of tools of excel
Created by jutao on 2019/8/27
E-mail jun.tao@nokia-sbell.com

"""
import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import datetime
import re

def date_diff_complex():

	file = input("excel to process: ")

	wb = xl.load_workbook(file)
	sheet = wb['Sheet1']

	wb_temp = xl.Workbook()
	sheet_temp = wb_temp.active

	max_row_number = sheet.max_row
	max_column_number = sheet.max_column

	reported_date_column = 1
	closed_date_column = 1

	while sheet.cell(1, reported_date_column).value != 'Reported Date':
		reported_date_column += 1

	while sheet.cell(1, closed_date_column).value != 'State Changed to Closed':
		closed_date_column += 1

	shortest_fix = 1000
	longest_fix = 0
	shortest_fix_row = 1
	longest_fix_row = 1
	fixtime_total = 0
	i = 1

	for row_number in range(2, max_row_number):
		# read date of pronto reported and closed
		reported_date = sheet.cell(row=row_number, column=reported_date_column).value
		closed_date = sheet.cell(row=row_number, column=closed_date_column).value

		if closed_date != '< empty >':
		# covert data form from date.month.year to year/month/date
			(re_date, re_month, re_year) = str(reported_date).split('.')
			reported_date = re_year + '/' + re_month + '/' + re_date
			(cl_date, cl_month, cl_year) = str(closed_date).split('.')
			closed_date = cl_year + '/' + cl_month + '/' + cl_date

			# assign date with new form to cell
			sheet_temp.cell(row=i, column=1).value = reported_date
			sheet_temp.cell(row=i, column=2).value = closed_date

			# caculate date difference
			x = 'A'+str(i)
			y = 'B'+str(i)
			z = '='+y+'-'+x
			sheet_temp.cell(row=i, column=3).value = z

			i += 1

	wb_temp.save('d:\\temp.xlsx')


	wb_temp = xl.load_workbook('d:\\temp.xlsx', data_only=True)
	sheet_temp = wb_temp['Sheet']

	for j in range(1, i):
		fixtime = sheet_temp.cell(row=j, column=3).value
		fixtime_total = fixtime_total + int(fixtime)

		if shortest_fix > int(fixtime):
			shortest_fix = int(fixtime)

		if longest_fix < int(fixtime):
			longest_fix = int(fixtime)


	fixtime_average = fixtime_total/(i-1)

	print('the shortest fix time: ' + str(shortest_fix))
	print('\nthe longest fix time ' + str(longest_fix))
	print('\naverage fix time: ' + str(fixtime_average))

# statistic pronto average fix time, pronto of shortest fix time, and pronto of longest fix time
def date_diff(file):
	wb = xl.load_workbook(file)
	ws = wb['Sheet1']
	column_list = []

	# list of title of each column
	for i in range(0,ws.max_column-1):
		column_list.append(ws.cell(row=1, column=i+1).value)

	reported_column_id = column_list.index('Reported Date') + 1
	closed_column_id = column_list.index('State Changed to Closed') + 1

	date_diff_sum = datetime.timedelta(0)
	shortest_fix = datetime.timedelta(1000)
	longest_fix = datetime.timedelta(0)
	shortest_fix_id = 0
	longest_fix_id = 0

	for i in range(2, ws.max_row+1):
		if str(ws.cell(row=i, column=closed_column_id).value) != '< empty >':
			try:
				(reported_date, reported_month, reported_year) = str(ws.cell(row=i, column=reported_column_id).value).split('.')
			except Exception as e:
				print('error happened ' + e.__class__.__name__)
			reported_time = datetime.date(int(reported_year), int(reported_month), int(reported_date))
			try:
				(closed_date, closed_month, closed_year) = str(ws.cell(row=i, column=closed_column_id).value).split('.')
			except Exception as e:
				print('error happened ' + e.__class__.__name__)
			closed_time = datetime.date(int(closed_year), int(closed_month), int(closed_date))
			date_diff = closed_time - reported_time
			date_diff_sum = date_diff_sum + date_diff

			if shortest_fix > date_diff:
				shortest_fix = date_diff
				shortest_fix_id = i

			if longest_fix < date_diff:
				longest_fix = date_diff
				longest_fix_id = i

	date_diff_ave = date_diff_sum/(ws.max_row-1)


	print('pronto average fix time: '+str(date_diff_ave))
	print(str(ws.cell(row=shortest_fix_id, column=1).value)+' has the fasted fix: '+str(shortest_fix))
	print(str(ws.cell(row=longest_fix_id, column=1).value)+' has the longest fix: '+str(longest_fix))
	
# change date format from day.month.year (31.1.2019) to year.month.day (2019/1/31)
def date_format_normalization(source, destination):
	wb = xl.load_workbook(source)
	ws = wb['Sheet1']
	column_list = []

	# list of title of each column
	for i in range(0,ws.max_column-1):
		column_list.append(ws.cell(row=1, column=i+1).value)

	reported_column_id = column_list.index('Reported Date') + 1
	closed_column_id = column_list.index('State Changed to Closed') + 1

	for i in range(2, ws.max_row+1):
		try:
			(reported_date, reported_month, reported_year) = str(ws.cell(row=i, column=reported_column_id).value).split('.')
		except Exception as e:
			print('error happened ' + e.__class__.__name__)
		ws.cell(row=i, column=reported_column_id).value = reported_year + '/' + reported_month + '/' + reported_date
		if str(ws.cell(row=i, column=closed_column_id).value) != '< empty >':
			(closed_date, closed_month, closed_year) = str(ws.cell(row=i, column=closed_column_id).value).split('.')
			ws.cell(row=i, column=closed_column_id).value = closed_year + '/' + closed_month + '/' + closed_date

	wb.save(destination)


# statistics pronto number proposed by each tester
def pr_num_per_tester(source, destination, stat_period_from, stat_period_to):
	wb = xl.load_workbook(source)
	ws = wb['Sheet1']
	column_list = []
	new_sheet_name = stat_period_to

	# add titles in row=1 into one list
	for i in range(0,ws.max_column-1):
		column_list.append(ws.cell(row=1, column=i+1).value)

	# filter out year, month, date in 'reported date' column and 'state changed to closed' column and create data of datetime type
	(year, month, date) = stat_period_from.split(',')
	stat_period_from = datetime.date(int(year), int(month), int(date))
	(year, month, date) = stat_period_to.split(',')
	stat_period_to = datetime.date(int(year), int(month), int(date))

	# find out which columns are about 'author' and 'reported date'
	author_column_id = column_list.index('Author') + 1
	report_date_column_id = column_list.index('Reported Date') + 1

	#create one set to store tester name, the reason to use set is duplicated items can be removed
	author_set = set()

	# add tester name into set, and via re to remove (Nokia - CN/Hangzhou), number in name and so on information in
	# tester name. e.g.Liu, Chao 5. (Nokia - CN/Hangzhou) is simplified to Liu, Chao
	for i in range(2, ws.max_row+1):
		temp_author = ws.cell(row=i, column=author_column_id).value
		temp_author = re.sub(r'\(.*$',"",temp_author)
		temp_author = re.sub(r'[\S]?\d.', "", temp_author)
		author_set.add(temp_author.strip())

	# create one list to store tester name, it's easy to use items in list as key in dictionary
	author_list = list(author_set)

	# create one dictionary to store tester name, it's easy to show tester and corresponding pronto number
	# like {'liu chao':100,'deng zhaori':90, 'liu zhiqiang':80}
	author_pr_dic = {}

	for i in range(0, len(author_list)):
		author_pr_dic[author_list[i]] = 0

	for i in range(2, ws.max_row+1):
		for j in range(0, len(author_list)):
			author_in_dic = author_list[j]
			author_in_cell = ws.cell(row=i, column=author_column_id).value
			try:
				(year, month, date) = (ws.cell(row=i, column=report_date_column_id).value).split('/')
			except Exception as e:
				print('Error happened ' + e.__class__.__name__)
			reported_date = datetime.date(int(year), int(month), int(date))
			# judge the pronto is statistics between defined start and end time
			if author_in_cell.find(author_in_dic) == 0 and stat_period_from <= reported_date and stat_period_to >= reported_date:
				author_pr_dic[author_list[j]] += 1

	# create one new sheet to show tester and corresponding pronto number information
	ws1 = wb.create_sheet(new_sheet_name)

 	# show tester and pronto number in new created sheet
	ws1['A1'] = 'tester'
	ws1['B1'] = 'pronto number'

	i = 2
	for key in author_pr_dic:
		ws1.cell(row=i, column=1).value = key
		ws1.cell(row=i, column=2).value = author_pr_dic[key]
		i += 1

	# create bar chart
	chart=xl.chart.BarChart()
	chart.type = "col"
	chart.style = 10
	chart.title = "Bar Chart"
	chart.y_axis.title = 'pronto number'
	chart.x_axis.title = 'tester'

	# define reference data for creating chart, use min-max row-column id to define the scope
	data = xl.chart.Reference(ws1, min_col=2, min_row=2, max_row=len(author_pr_dic))
	cats = xl.chart.Reference(ws1, min_col=1, min_row=2, max_row=len(author_pr_dic))
	chart.add_data(data, titles_from_data=True)
	chart.set_categories(cats)
	chart.shape = 4
	ws.add_chart(chart, "E10")

	wb.save(destination)
	wb.close()



# statistics pronto number of each feature
def pr_num_per_feature(source, destination):
	wb = xl.load_workbook(source)
	ws = wb['Sheet1']

	# create one list to involve all titles for each column
	column_list = []
	for i in range(0,ws.max_column-1):
		column_list.append(ws.cell(row=1, column=i+1).value)

	# find out the column id for 'Feature'
	feature_column_id = column_list.index('Feature') + 1

	# create one set to include all features, duplicated items in set can be removed automatically
	feature_set = set()

	for i in range(2, ws.max_row+1):
		feature_set.add(ws.cell(row=i, column=feature_column_id).value)

	# create one list to include all features, due to items in set are orderless
	feature_list = list(feature_set)

	# create one dictionary, key is 'feature', value is pronto number and initialized to zero
	feature_pr_dic = {}
	for i in range(0, len(feature_list)):
		feature_pr_dic[feature_list[i]] = 0

	# find out pronto number for each feature
	for i in range(2, ws.max_row+1):
		feature_pr_dic[ws.cell(row=i, column=feature_column_id).value] += 1

	# create one new sheet to show feature and corresponding pronto number in two column
	ws1 = wb.create_sheet('pr_num_per_feature')

	ws1['A1'] = 'feature'
	ws1['B1'] = 'pronto number'

	i = 2
	for key in feature_pr_dic:
		ws1.cell(row=i, column=1).value = key
		ws1.cell(row=i, column=2).value = feature_pr_dic[key]
		i += 1

	# create bar chart
	chart=xl.chart.BarChart()
	chart.type = "col"
	chart.style = 10
	chart.title = "Bar Chart"
	chart.y_axis.title = 'pronto number'
	chart.x_axis.title = 'feature'

	# define reference data for creating chart, use min-max row-column id to define the scope
	data = xl.chart.Reference(ws1, min_col=2, min_row=2, max_row=len(feature_pr_dic))
	cats = xl.chart.Reference(ws1, min_col=1, min_row=2, max_row=len(feature_pr_dic))
	chart.add_data(data, titles_from_data=True)
	chart.set_categories(cats)
	chart.shape = 4
	ws.add_chart(chart, "E10")

	wb.save(destination)
	wb.close()



# check cell in source excel if is in destination excel
def cell_in_another_excel(source, destination):
	wb_source = xl.load_workbook(source)
	ws_source = wb_source['Sheet1']

	wb_destination = xl.load_workbook(destination)
	ws_destination = wb_destination['Sheet1']

	column_in_source = int(input('column id in source excel to check: '))
	column_in_destination = int(input('column id in destination excel to compare: '))

	for i in range(2, ws_source.max_row+1):
		cell_to_check = ws_source.cell(row=i, column=column_in_source).value

		for j in range(2, ws_destination.max_row+1):
			if ws_destination.cell(row=j, column=column_in_destination) == cell_to_check:
				ws_source.cell(row=i, column=ws_source.max_column+1).value = 'Y'
				break


# check cell value in source file if is also in destination file
def  cell_value_also_in_another_excel(source, destination, output):
	wb_source = xl.load_workbook(source)
	ws_source = wb_source['Sheet1']

	wb_destination = xl.load_workbook(destination)
	ws_destination = wb_destination['Sheet1']

	column_in_source = int(input('column id in source excel to check: '))
	column_in_destination = int(input('column id in destination excel to compare: '))

	redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

	for i in range(2, ws_source.max_row+1):
		cell_to_check = ws_source.cell(row=i, column=column_in_source).value

		for j in range(2, ws_destination.max_row+1):
			cell_to_compare = ws_destination.cell(row=j, column=column_in_destination).value
			if  cell_to_compare == cell_to_check:
				row_to_lable = 'A' + str(i)
				# lable accurate cell (which value is same as that in destination file with red color
				ws_source[row_to_lable].fill = redFill
				break

	wb_source.save(output)


# check same case in two columns and mark both red color
def  same_case(file):
	wb = xl.load_workbook(file)
	ws = wb['Sheet1']

	redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

	for i in range(1, ws.max_row+1):
		cell_to_check = re.sub('\[.*\]', '', str(ws.cell(row=i, column=1).value))

		for j in range(1, ws.max_row+1):
			cell_to_compare = re.sub('\[.*\]', '', str(ws.cell(row=j, column=2).value))
			if (cell_to_check.upper()).lstrip() == (cell_to_compare.upper()).lstrip():
				print(cell_to_check)
				check_row_to_lable = 'A' + str(i)
				compare_row_to_lable = 'B' + str(j)
				ws[check_row_to_lable].fill = redFill
				ws[compare_row_to_lable].fill = redFill
				break

	wb.save(file)


def  simplify_cell_content(file):
	"""
	original cell content: 2021-07-01T21:12:19.643000
	simplified cell content: 2021-07-01
	create new column to save simplified cell content
	"""
	wb = xl.load_workbook(file)
	ws = wb['Sheet1']

	# from which letter to cut following part, in this case, delimeter is 'T'
	delimeter = input('from where to cut:')

	for i in range(1, ws.max_row+1):
		original_cell_content = str(ws.cell(row=i, column=1).value)
		from_where_to_cut = original_cell_content.index(delimeter)
		simplified_cell_content = original_cell_content[:from_where_to_cut]
		print('{0} -> {1}'.format(original_cell_content, simplified_cell_content))
		ws.cell(row=i, column=2).value = simplified_cell_content

	wb.save(file)


if	__name__ == '__main__':
	file = 'D:\\20210919.xlsx'
	simplify_cell_content(file)



