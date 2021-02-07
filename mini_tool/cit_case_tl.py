"""
statistics which cit cases run on which test line
below is report example:
Name	                                                            Duration (diff)	            test line
vDU Snapshot Collection To LSS NMS 5GC001662-I	                  0:19:29.279 (+0:04:32.406)	CLOUD_5G_AIC-7_5_UTE5GAIC001T094-Durian
IPv4 On Demand Path MTU Discovery With WebEM For vDU 5GC001995-C	0:31:33.791 (+0:31:33.791)	CLOUD_5G_AIC-7_5_UTE5GAIC001T094-Durian
"""

from bs4 import BeautifulSoup
import requests, sys, webbrowser, bs4
import pandas as pd
import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font, Border
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
# from collections import Counter


# flag, if the first dataframe written in csv, if yes, keep the column name; if not, remove the column name
is_first_record_in_csv = True

def get_url_of_robot_result(test_line_list):
	"""
	find the url of 'robot results' in jinkens which contains cit case list
	"""
	for test_line in test_line_list:
		url=r'http://10.109.2.191:8888/job/' + test_line
		res=requests.get(url)
		try:
			res.raise_for_status()
		except Exception as exc:
			print('there is a problem:%s'%(exc))

		bs_res = BeautifulSoup(res.text, 'html.parser')

		# find the link of 'robot results' in http://10.109.2.191:8888/job/CLOUD_5G_AIC-7_5_UTE5GAIC001T066-Matador/
		robot_result_url_postfix = bs_res.find('a', attrs={'class':'task-link','title':"Robot Results"}).attrs['href']
		print(robot_result_url_postfix)

		# complement the url of 'robot results'
		robot_result_url = r'http://10.109.2.191:8888' + robot_result_url_postfix
		print(robot_result_url)
		webbrowser.open(robot_result_url)

		output_cit_case_to_csv(robot_result_url, test_line)


def output_cit_case_to_csv(url, test_line):
	"""
	read cit case table and write to csv file
	"""
	df = pd.DataFrame()
	try:
		# ignore_index 默认为 False ，当设为 ignore_index=True 时，新 df 将不会使用拼接成员
		# df 的 index，而是重新生成一个从 0 开始的 index 值
		df = df.append(pd.read_html(url)[-1], ignore_index=False)
	except Exception as exc:
		print('there is a problem:%s'%(exc))
	else:
		df['TL'] = test_line
		df['TL'][0] = 'test line'

		global is_first_record_in_csv
		print(is_first_record_in_csv)
		if not is_first_record_in_csv:
			# DataFrame.drop(labels=None,axis=0, index=None, columns=None, inplace=False)
			# labels 就是要删除的行列的名字，用列表给定
			# axis 默认为0，指删除行，因此删除columns时要指定axis=1；
			# index 直接指定要删除的行
			# columns 直接指定要删除的列
			# inplace=False，默认该删除操作不改变原数据，而是返回一个执行删除操作后的新dataframe；
			# inplace=True，则会直接在原数据上进行删除操作，删除后无法返回
			df.drop(index=0, axis=0, inplace=True)

		print(df)
		# only write columns:0,3,4 of original table, mode is append which means append new contents in old csv file
		df.to_csv('cit_case.csv', columns=[df.columns[0], df.columns[3], df.columns[4]], header=0, index=False, mode='a')
		is_first_record_in_csv = False


def csv_covert_to_xlsx():
	"""
	covert csv format to xlsx format
	"""
	csv = pd.read_csv('cit_case.csv', encoding='utf-8')
	csv.to_excel('cit_case.xlsx', sheet_name='raw_data', index=False)


def statistics_excel():
	"""
	statistic below data in csv file:
	1. duplicated cases and highlighted with red color
	2. how many cit cases run on each test line
	3. total run duration on each test line
	...

	"""
	file = 'cit_case.xlsx'
	wb = xl.load_workbook(file)
	ws = wb['raw_data']

	# add case name in one empty list i.e. case_list
	case_list = []
	for i in range(1, ws.max_row+1):
		case_list.append(ws.cell(row=i, column=1).value)

	print(case_list)

	# find the duplicated item in case_list and marked as red color
	redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

	for item in case_list:
		# find items which has more than one occurrence
		if case_list.count(item) > 1:
			print(item)
			duplicated_item_row = case_list.index(item)
			duplicated_item_cell = 'A' + str(duplicated_item_row+1)
			ws[duplicated_item_cell].fill = redFill

	# add test line to one empty list i.e. test_line_list
	test_line_list = []
	for i in range(2, ws.max_row+1):
		test_line_list.append(ws.cell(row=i, column=3).value)

	# create one dictionary, key is test line name, value is how many cases run on this test line
	test_line_case_count_dic = dict()
	test_line_set = set(test_line_list)

	for test_line in test_line_set:
		test_line_case_count_dic[test_line] = test_line_list.count(test_line)

	print(test_line_case_count_dic)

	"""
	create new sheet to summarize
	1. how many cases run on each test line
	2. how long all cases run on each test line
	"""
	ws2 = wb.create_sheet('overview')
	ws2.cell(row=1, column=1).value = 'test line'
	ws2.cell(row=1, column=2).value = 'case count'
	ws2.cell(row=1, column=3).value = 'duration'
	j = 2
	for key, value in test_line_case_count_dic.items():
		ws2.cell(row=j, column=1).value = key
		ws2.cell(row=j, column=2).value = value
		j += 1

	wb.save(file)


def send_report_mail():
	"""
	send reminder mail to case responsible person once one pronto was closed
	mail format is below
	mail title : case re-test reminder due to pronto closed
	from : jun.tao@nokia-sbell.com
	to :
	content:
	Hi, <responsible person name>
	<PRONTO ID> was just closed, please re-test below case asap, thanks!
	<case name>
	"""
	mail_host = "mail.int.nokia-sbell.com"

	sender = "jun.tao@nokia-sbell.com"
	receiver = "jun.tao@nokia-sbell.com"

	message = MIMEMultipart()
	message['From'] = Header('jun.tao@nokia-sbell.com', 'utf-8')
	message['To'] = Header('jun.tao@nokia-sbell.com', 'utf-8')
	subject = 'CIT test line load report'
	message['Subject'] = Header(subject, 'utf-8')
	# mail body
	message.attach(MIMEText('CIT test line load report\nnote:cases with RED color means duplicated on different TL', 'plain', 'utf-8'))

	# construct attachment
	attachment = MIMEText(open('cit_case.xlsx', 'rb').read(), 'base64', 'utf-8')
	attachment["content-type"] = 'application/octect-stream'
	attachment["content-Disposition"] = 'attachment; filename="cit_case.xlsx"'
	message.attach(attachment)

	try:
		smtp_obj = smtplib.SMTP()
		smtp_obj.connect(mail_host)
		smtp_obj.sendmail(sender, receiver, message.as_string())
		print('sending succeed')
		return 'ok'
	except smtplib.SMTPEexception:
		print('sending failed')
		return 'nok'


def main():
	test_line_list = ['CLOUD_5G_AIC-7_5_UTE5GAIC001T094-Durian',
				 'CLOUD_5G_AIC-7_5_UTE5GAIC001T030-Apple-5324',
				 'CLOUD_5G_AIC-7_5_UTE5GAIC001T066-Matador',]
	get_url_of_robot_result(test_line_list)
	csv_covert_to_xlsx()
	statistics_excel()
	print('please find cit_case.xlsx in D:\python\helloworld\web')


if __name__ == '__main__':
	send_report_mail()
