"""
check pronto status in pronto statistics excel, if it's closed, send remind mail to responsible tester
"""

import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
import time
from datetime import date
import re
import requests, sys, webbrowser, bs4
import smtplib
from email.mime.text import MIMEText
from email.header import Header


mailbox_dic = {'Shan Zhijiang':'zhijiang.shan@nokia-sbell.com',
			   'Deng Zhaori':'zhaori.deng@nokia-sbell.com',
			   'Ni Gang':'gang.ni@nokia-sbell.com',
			   'Qiu Xiubin':'xiubin.qiu@nokia-sbell.com',
			   'Qu Hefeng':'hefeng-cookie.qu@nokia-sbell.com',
			   'Rao Hongping':'hongping.rao@nokia-sbell.com',
			   'Ren Shubin':'shubin.ren@nokia-sbell.com',
			   'Lin Peng':'peng.lin@nokia-sbell.com',
			   'Liu Chao':'chao.5.liu@nokia-sbell.com',
			   'Feng Rita':'rita.feng@nokia-sbell.com',
			   'Liu Zhiqiang':'zhiqiang.q2.liu@nokia-sbell.com',
			   'Chen Zhong':'zhong.1.chen@nokia-sbell.com',
			   'Cao Chris':'chris.l.cao@nokia-sbell.com',
			   'Xu Eahsuan':'eahsuan.xuan@nokia-sbell.com',
			   'Wang Lijuan':'lijuan.1.wang@nokia-sbell.com',
			   'Su Winnie':'winnie.su@nokia-sbell.com',
			   'Zhao Ling':'ling.2.zhao@nokia-sbell.com',
			   'Xu Huilin':'huilin.xu@nokia-sbell.com',
			   'Chen Wei':'wei.7.chen@nokia-sbell.com',
			   'Tao Jun':'jun.tao@nokia-sbell.com'}


def get_pronto_case_tester(source_file):
	"""
	check column 'Fault Report ID' in source file, if there is one pronto id, create one dictionary
	{'PR ID':'value in column Fault Report ID,
 	'Case name':value in column Test Instance Name,
 	'tester':value in column Responsible Tester}
	at last,return one list of dictionary.
	"""
	wb = xl.load_workbook(source_file)
	ws = wb['Sheet1']

	pronto_case_tester_list = []

	for i in range(1, ws.max_row+1):
		# read information of column 'Fault Report ID'
		pronto_info = ws.cell(row=i, column=column_index_from_string('G')).value
		# if pronto id is not empty, continue to proceed
		if pronto_info.strip() != '':
			print(pronto_info)
			# separate each pronto from 'PR00001;PR00002' such kinds of string
			pronto_list = pronto_info.split(';')
			for pronto_id in pronto_list:
				pronto_case_tester_dic = {}
				# check if begin from 'P' due to Jira or comments maybe also in column 'Fault Report ID'
				if pronto_id[0] == 'P':
					pronto_status = get_pronto_status(pronto_id)
					print(pronto_status)
					if pronto_status == 'Closed':
						# remove prefix '[1][1.0]' in case name
						case_name = (re.sub('\[.*\]', '', str(ws.cell(row=i, column=column_index_from_string('B')).value))).strip()
						# remove '(xxx)' and '.number' in tester name
						tester = ws.cell(row=i, column=column_index_from_string('AB')).value
						tester = re.sub(r'\(.*$',"",tester)
						tester = (re.sub(r'[\S]?\d.', "", tester)).strip()
						pronto_case_tester_dic['case'] = case_name
						pronto_case_tester_dic['pronto'] = pronto_id
						pronto_case_tester_dic['tester'] = tester

						# added dic which meet the requirement that pronto already closed into list
						pronto_case_tester_list.append(pronto_case_tester_dic)

#	for pronto_case_tester_item in pronto_case_tester_list:
#		print(pronto_case_tester_item)

	return pronto_case_tester_list


def get_pronto_status(pronto_id):
	"""
	check pronto status via reading field 'state' in pronto tool
	"""
	# url in pronto too
	url=r'https://pronto.int.net.nokia.com/prontoapi/rest/api/1/problemReport/{}'.format(str(pronto_id))
	res=requests.get(url,auth=('horao', 'aaaaaa1!'))
	try:
		res.raise_for_status()
	except Exception as e:
		print("enter pronto page failed: id={},reason={}".format(pronto_id,str(e)))
	content = res.json()
	print(content)
	try:
		status = content['state']
		return status
	except Exception as e:
		print('get status failed: reason={}'.format(str(e)))
		return 'get_status_failed'


def send_reminder_mail(pronto_case_tester_dic):
	"""
	send reminder mail to case responsible person once one pronto was closed
	mail format is below
	mail title : case re-test reminder due to pronto closed
	from : jun.tao@nokia-sbell.com
	to : mail box of responsible person
	content:
	Hi, <responsible person name>
	<PRONTO ID> was just closed, please re-test below case asap, thanks!
	<case name>
	"""
	mail_host = "mail.int.nokia-sbell.com"

	sender = "jun.tao@nokia-sbell.com"
	receiver = mailbox_dic[pronto_case_tester_dic['tester']]

	mail_body = 'Hi,' + pronto_case_tester_dic['tester'] + '\n' + pronto_case_tester_dic['pronto'] + \
				' was just closed, please re-test below case asap, thanks! \n' + pronto_case_tester_dic['case']
	message = MIMEText(mail_body, 'plain', 'utf-8')
	message['From'] = Header('jun.tao@nokia-sbell.com', 'utf-8')
	message['To'] = Header('jun.tao@nokia-sbell.com', 'utf-8')

	subject = 'case re-test reminder due to pronto closed'
	message['Subject'] = Header(subject, 'utf-8')

	try:
		smtp_obj = smtplib.SMTP()
		smtp_obj.connect(mail_host)
		smtp_obj.sendmail(sender, receiver, message.as_string())
		print('sending succeed')
		return 'ok'
	except smtplib.SMTPEexception:
		print('sending failed')
		return 'nok'


def history_log(destination_file, history_log_list):
	"""
	record each item if mail sending succeed.
	record format:
	case name              pronto id         tester      remind result      date
	TRS_IP_connectivity    PR00001           Tao Jun     ok                 20/1/2021
	"""
	wb = xl.Workbook()
	ws = wb.active

	ws['A1'] = 'case name'
	ws['B1'] = 'pronto id'
	ws['C1'] = 'tester'
	ws['D1'] = 'remind result'
	ws['E1'] = 'date'

	length_of_history_log_list = len(history_log_list)

	for i in range(1, length_of_history_log_list):
		ws.cell(row=i+1, column=1).value = history_log_list[i]['case']
		ws.cell(row=i+1, column=2).value = history_log_list[i]['pronto']
		ws.cell(row=i+1, column=3).value = history_log_list[i]['tester']
		ws.cell(row=i+1, column=4).value = history_log_list[i]['remind result']
		ws.cell(row=i+1, column=5).value = history_log_list[i]['date']

	wb.save(destination_file)
	wb.close()


if __name__ == '__main__':
	source_file=r'd:\feature_pronto_2.xlsx'
	pronto_case_tester_list = get_pronto_case_tester(source_file)
	# create a new list of dictionary and add key 'remind result' and 'date'
	history_log_list = pronto_case_tester_list.copy()

	current_date=date.today()
	ticks = str(time.time())
	destination_file = 'd:\log-' + ticks + '.xlsx'

	for history_log_item in history_log_list:
		remind_result = send_reminder_mail(history_log_item)
		history_log_item['remind result'] = remind_result
		# translate date format to '2021-1-20'
		history_log_item['date'] = current_date.isoformat()

	history_log(destination_file, history_log_list)
